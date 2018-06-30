import openpyxl
import json

output_hdr = [  "Campaign",
                "Ad Group",
                "Labels",
                "Keyword",
                "Criterion Type",
                "Bid Strategy Type",
                "Bid Strategy Name",
                "Max CPC",
                "Max CPV",
                "First page bid",
                "Top of page bid",
                "First position bid",
                "Quality score",
                "Landing page experience",
                "Expected CTR",
                "Ad relevance",
                "Destination URL",
                "Final URL",
                "Final mobile URL",
                "Tracking template",
                "Custom parameters",
                "Campaign Status",
                "Ad Group Status",
                "Status",
                "Approval Status",
                "Comment",
                "CPA Bid",
                "Display Network Custom Bid Type",
                "Targeting optimization",
                "Ad Group Type",
                "Flexible Reach",
                "Max CPM",
                "Campaign Daily Budget",
                "Campaign Type",
                "Networks",
                "Languages",
                "Bid Strategy Type",
                "Bid Strategy Name",
                "Enhanced CPC",
                "Ad rotation",
                "Delivery method",
                "Targeting method",
                "Exclusion method",
                "DSA targeting source",
                "DSA page feeds",
                "Flexible Reach",
                "Headline 1",
                "Headline 2",
                "Description",
                "Path 1",
                "Path 2",
                "Final URL"
            ]

output_map = {}
i = 0
for entry in output_hdr:
    output_map[entry] = i
    i += 1

output_template = ["" for i in range(len(output_hdr))]

def readInputData(filename="conteudo2.xlsx"):
    content = {}
    try:
        wb = openpyxl.load_workbook(filename)
    except IOError:
        print("Erro ao abrir o arquivo de entrada %s.")
        return None

    for sheet_name in wb.get_sheet_names():
        ws = wb[sheet_name]
        content[sheet_name] = []
        for row in range(2, ws.max_row + 1):
            line = []
            for col in range(1, ws.max_column + 1):
                value = ws.cell(column=col, row=row).value
                line.append(value)
            content[sheet_name].append(line)
        if sheet_name == "Campanha" or sheet_name == "Seeds":
            content[sheet_name] = transpose(content[sheet_name])

    return content

def writeKeywordEntry(c, a, k, output):
    to_output = [k for k in output_template]
    to_output[output_map["First page bid"]] = "0.00"
    to_output[output_map["Top of page bid"]] = "0.00"
    to_output[output_map["First position bid"]] = "0.00"
    to_output[output_map["Campaign Status"]] = "Enabled"
    to_output[output_map["Ad Group Status"]] = "Enabled"
    to_output[output_map["Status"]] = "Enabled"
    to_output[output_map["Campaign"]]=c
    to_output[output_map["Ad Group"]]=a
    to_output[output_map["Keyword"]]=k
    if "-butterfly" in c:
        to_output[output_map["Criterion Type"]]="Broad"
        writeCSVEntry(output, to_output)
        to_output[output_map["Criterion Type"]]="Negative Exact"
        writeCSVEntry(output, to_output)
        to_output[output_map["Criterion Type"]]="Negative Phrase"
        writeCSVEntry(output, to_output)
    else:
        to_output[output_map["Criterion Type"]]="Phrase"
        writeCSVEntry(output, to_output)
        to_output[output_map["Criterion Type"]]="Exact"
        writeCSVEntry(output, to_output)

def writeAdEntry(c, a, ad, output):
    to_output = [k for k in output_template]
    to_output[output_map["Headline 1"]] = ad[0]
    to_output[output_map["Headline 2"]] = ad[1]
    to_output[output_map["Description"]] = ad[2]
    to_output[output_map["Path 1"]] = ad[3]
    to_output[output_map["Path 2"]] = ad[4]
    to_output[output_map["Final URL"]] = ad[5]
    to_output[output_map["Campaign"]]=c
    to_output[output_map["Ad Group"]]=a
    writeCSVEntry(output, to_output)


def writeCSVEntry(oFile, entry):
    to_write = []
    for e in entry:
        if "," in e:
            to_write.append('"' + e + '"')
        else:
            to_write.append(e)
    strEntry = ",".join(to_write) + "\r\n"
    oFile.write(strEntry)

def transpose(M):
    MT = []
    for j in range(len(M[0])):
        l = []
        for i in range(len(M)):
            if M[i][j] is not None and M[i][j] != "":
                l.append(M[i][j])
        if len(l) == 0:
            l.append("")
        MT.append(l)
    return MT

def loadSKAGControl(c):
    try:
        filename = "SKAGControl/SKAGControl_" + c + ".json"
        controlFile = open(filename, "r", encoding="utf-16-le")
    except IOError:
        input("Criando arquivo de controle de SKAGs %s.\r\nPressione ENTER para continuar." %filename)
        dataModel = {"filename":filename,"versions": [{"date": "","adgroups": {}}]}
        saveSKAGControl(dataModel)
        return loadSKAGControl(c)

    try:
        skagControl = json.load(controlFile, encoding="utf-16-le")
    except ValueError:
        print("Erro ao parsear JSON de controle de SKAGs")
        return None

    controlFile.close()
    return skagControl

def saveSKAGControl(skagObject):
    try:
        controlFile = open(skagObject["filename"], "w", encoding="utf-16-le")
    except:
        print("Erro ao carregar arquivo de controle de SKAGs %s (saída)" %skagObject["filename"])
        return None

    json.dump(skagObject, controlFile, indent=4)

    controlFile.close()

def initConectors():
    try:
        conectorsFile = open("conectores.txt", "r", encoding="utf-16-le")
    except:
        print("Erro ao carregar arquivo de conectores")
        input()
        return []

    conectors = []
    for line in conectorsFile:
        conectors.append(line.replace("\r", "").replace("\n", ""))

    conectorsFile.close()

    return conectors
